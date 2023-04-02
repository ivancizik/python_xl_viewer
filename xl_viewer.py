import openpyxl
import os
import sys
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import column_index_from_string
from prettytable import PrettyTable


class dataTable:
    def __init__(self):
        self.table = PrettyTable()

    def add_columns(self, column_no: int):
        self.reset()
        columns = [" "]
        for i in range(1, column_no):
            columns.insert(i, get_column_letter(i))
        self.table.field_names = columns

    def add_row(self, rows: list):
        self.table.add_row(rows)

    def get_table(self):
        return self.table

    def reset(self):
        self.table = PrettyTable()


class excelFile:
    def __init__(self):
        self.sheet = None
        self.wb = None
        self.column_no = 0

    def open_file(self, file: str):
        try:
            self.wb = openpyxl.load_workbook(file)
            self.sheet = self.wb.active
            self.column_no = int(self.sheet.max_column + 1)
        except Exception as e:
            print("Error opening file: ", xl_file)
            print(str(e))
            quit()

    def get_column_no(self):
        return self.column_no

    def get_row_no(self):
        return self.sheet.max_row

    def get_row(self, row_no: int):
        row = [row_no]
        for i in range(1, self.column_no):
            if self.sheet.cell(row=row_no, column=i).value in ["None", None]:
                row.insert(i, "")
            else:
                row.insert(i, self.sheet.cell(row=row_no, column=i).value)
        return row

    def change_active_sheet(self, sheet: str):
        if sheet in self.wb.sheetnames:
            self.sheet = self.wb[sheet]
            self.column_no = int(self.sheet.max_column + 1)
        else:
            print("Invalid sheet name")
            return 0

    def remove_sheet(self, sheet: str):
        self.wb.remove(self.wb[sheet])

    def generate_search_results(self, xl_search_term: str, column: str):
        self.wb.create_sheet("Results")
        xl_search_term_offset = 0
        if xl_search_term[-2:] == "-h":
            xl_search_term_offset = 3
            for row in self.sheet.iter_rows(min_row=1, max_row=2):
                self.wb["Results"].append((c.value for c in row))
        for row in self.sheet.iter_rows():
            search_match = False
            cell_value = str(row[column_index_from_string(column[0]) - 1].value)

            if cell_value == "None":
                pass

            if xl_search_term[0] == "*" and xl_search_term[-1 - xl_search_term_offset] == "*":
                search_match = xl_search_term.split("*")[1].lower() in cell_value.lower()
            else:
                search_match = xl_search_term.split(" -")[0].lower() == cell_value.lower()

            if xl_search_term[0] == "*":
                search_match = cell_value.lower().endswith(xl_search_term.split("*")[1].split(" -")[0].lower())

            if xl_search_term[-1 - xl_search_term_offset] == "*":
                search_match = cell_value.lower().startswith(xl_search_term.split("*")[0].lower())

            if search_match:
                self.wb["Results"].append((c.value for c in row))

        self.change_active_sheet("Results")


def generate_table():
    os.system('cls' if os.name == 'nt' else 'clear')
    t.add_columns(f.get_column_no())
    for i in range(1, f.get_row_no()):
        t.add_row(f.get_row(i))

    print(f.sheet)
    print(t.get_table())
    if "Results" in f.wb.sheetnames:
        f.remove_sheet("Results")
        print(f.wb.sheetnames)
        print("1 - Open sheet. 3 - Exit")
    else:
        print(f.wb.sheetnames)
        print("1 - Open sheet. 2 - Search. 3 - Exit")


def action(option: str):
    match option:
        case "1":
            while f.change_active_sheet(input("Open Sheet: ")) == 0:
                pass
            t.table.clear_rows()
            generate_table()

        case "2":
            print(
                '''
                Use the following search functions:
                string      = check is searched item is a match
                *string     = check is searched item ends with string
                string*     = check is searched item starts with string
                *string*    = check is searched item contains with string
                -h          = add headers to search results (first row is header)
                '''
            )
            f.generate_search_results(input("Search for: "), input("Select column: "))
            t.table.clear_rows()
            generate_table()

        case "3":
            os.system('cls' if os.name == 'nt' else 'clear')
            sys.exit()

        case _:
            print("Invalid action")


if __name__ == "__main__":
    f = excelFile()
    xl_file = "" if len(sys.argv) <= 1 else sys.argv[1]
    f.open_file(xl_file)
    t = dataTable()
    generate_table()

    while True:
        action(input("Select action: "))
